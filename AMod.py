# -*- coding: utf-8 -*-
from __future__ import unicode_literals
import time
import datetime
import Tkinter
from Tkinter import *
import ttk
from ttk import *
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import visa
import os
import tkMessageBox as mb
from PIL import Image, ImageTk


yellowFill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
rm = visa.ResourceManager()
rm1=rm.list_resources()
rmtip=list(rm1)
Cnanel = 0
Cnanel1 = 0
Save1 = ""
inst_1 = 0
inst_2 = 0
ws = 0
wb = 0
span = ["100Hz", "1kHz", "10kHz", "100kHz", "1MHz", "10MHz"]
xz = 0

def tip():
        global rm1
        global rmtip
        v=len(rm1)
        i=0
        while i < v:
                b1=rm1[i]
                index_b1=b1.find('0x0D0B') # product ID
                if index_b1 >= 0:
                        rmtip[i]='N9030A'
                    
                i=i+1
        combo1.config(values=rmtip)

class id_tip():
        def __init__(self, tipP, tipID, combo, x_combo ):
                self.tipP = tipP
                self.tipID = tipID
                self.combo = combo
                self.x_combo = x_combo
        def id_tip(self):
                global rm1
                global rmtip
                global e
                v=len(rmtip)
                i=0
                b1=(self.combo).get()
                index_b1=b1.find(self.tipP)
                if index_b1 >= 0:
                        while i < v:
                                b2=rm1[i]
                                index_b2=b2.find(self.tipID)
                                if index_b2 >= 0:
                                        (self.x_combo).set(rm1[i])
                                i=i+1
                
def click_connect():
        x = id_tip('N9030A','0x0D0B', combo1,e) #class id_tip
        x.id_tip()
        global Save1
        global inst_1
        today1 = today.strftime("%d.%m.%Y %H.%M.%S")
        port=combo1.get()
        port1 = port[0:4]
        if port1 == "USB0":
                inst_1 = rm.open_resource(e.get())
                data_1 = inst_1.query("*IDN?")
                a.set(data_1)
                lab1.config(text="ID Analyzer:")
                pb1.config(mode="indeterminate")
                a1=entry1.get()
                index_tip1=a1.find("N9030A")
                if index_tip1 >= 0:
                        text.config(state=NORMAL)
                        text.insert(END, "N9030A - готов\n")
                        text.see("end")
                        text.config(state=DISABLED)
                        but3.config(state=NORMAL)
                else:
                        text.config(state=NORMAL)
                        text.insert(END, "не поддерживается\n \n")
                        text.see("end")
                        text.config(state=DISABLED)       
                        
class ed_izm():
        def __init__(self, izm, Gx, Mx, kx, x, mx, mkx, nx):
                self.izm = izm
                self.Gx = Gx
                self.Mx = Mx
                self.kx = kx
                self.x = x
                self.mx = mx
                self.mkx = mkx
                self.nx = nx
        def ed_izm(self):
                global xz
                if self.izm >= 1000000000:
                        self.izm = self.izm / 1000000000
                        self.izm = str(self.izm) + self.Gx
                        xz = self.izm
                elif self.izm >= 1000000:
                        self.izm = self.izm / 1000000
                        self.izm = str(self.izm) + self.Mx
                        xz = self.izm
                elif self.izm >= 1000:
                        self.izm = self.izm / 1000
                        self.izm = str(self.izm) + self.kx
                        xz = self.izm
                elif self.izm >= 1:
                        self.izm = self.izm / 1
                        self.izm = str(self.izm) + self.x
                        xz = self.izm
                elif self.izm >= 0.001:
                        self.izm = self.izm * 1000
                        self.izm = str(self.izm) + self.mx
                        xz = self.izm
                elif self.izm >= 0.000001:
                        self.izm = self.izm * 1000000
                        self.izm = str(self.izm) + self.mkx
                        xz = self.izm
                elif self.izm >= 0.000000001:
                        self.izm = self.izm * 1000000000
                        self.izm = str(self.izm) + self.nx
                        xz = self.izm
                elif self.izm <= -1000000000:
                        self.izm = self.izm / 1000000000
                        self.izm = str(self.izm) + self.Gx
                        xz = self.izm
                elif self.izm <= -1000000:
                        self.izm = self.izm / 1000000
                        self.izm = str(self.izm) + self.Mx
                        xz = self.izm
                elif self.izm <= -1000:
                        self.izm = self.izm / 1000
                        self.izm = str(self.izm) + self.kx
                        xz = self.izm
                elif self.izm <= -1:
                        self.izm = self.izm / 1
                        self.izm = str(self.izm) + self.x
                        xz = self.izm
                elif self.izm <= -0.001:
                        self.izm = self.izm * 1000
                        self.izm = str(self.izm) + self.mx
                        xz = self.izm
                elif self.izm <= -0.000001:
                        self.izm = self.izm * 1000000
                        self.izm = str(self.izm) + self.mkx
                        xz = self.izm
                elif self.izm <= -0.000000001:
                        self.izm = self.izm * 1000000000
                        self.izm = str(self.izm) + self.nx
                        xz = self.izm       
           
def click_startAM():
        global span
        inst_1.write("FREQuency:TUNE:IMM")
        s=0
        T=(1*1000)+1
        while s<T: 
                time.sleep (0.1)
                pb1.step(1)
                root.update()
                s=s+100
        inst_1.write("FREQ:SPAN "+combo2.get())
        s=0
        T=(1*1000)+1
        while s<T: 
                time.sleep (0.1)
                pb1.step(1)
                root.update()
                s=s+100
        inst_1.write("CALC:MARK1:MAX")
        inst_1.write("CALC:MARK2:MAX")
        inst_1.write("CALC:MARK2:MAX:NEXT")
        #inst_1.write("CALC:MARK2:MAX:RIGH")
        inst_1.write("CALC:MARK3:MAX")
        #inst_1.write("CALC:MARK3:MAX:LEFT")
        inst_1.write("CALC:MARK3:MAX:NEXT")
        inst_1.write("CALC:MARK3:MAX:NEXT")
        inst_1.write("UNIT:POWer V")
        
        mark1 = inst_1.query("CALC:MARK1:Y?")
        mark1 = float(mark1)
        x = ed_izm(mark1, ' ГВ', ' МВ', ' кВ', ' В', ' мВ', ' мкВ', ' нВ')
        x.ed_izm()
        mark11.set(xz)
        #mark11.set("{0:.5f} мВ".format(mark1))
        mark1f = inst_1.query("CALC:MARK1:X?")
        mark1f = float(mark1f)
        x = ed_izm(mark1f, ' ГГц', ' МГц', ' кГц', ' Гц', ' мГц', ' мкГц', ' нГц')
        x.ed_izm()
        mark21.set(xz)
        #mark121f= mark1f/1000000000
        #mark21.set("{0:.6f} ГГц".format(mark121f))
        
        mark2 = inst_1.query("CALC:MARK2:Y?")
        mark2 = float(mark2)
        x = ed_izm(mark2, ' ГВ', ' МВ', ' кВ', ' В', ' мВ', ' мкВ', ' нВ')
        x.ed_izm()
        mark12.set(xz)
        #mark12.set("{0:.5f} мВ".format(mark2))
        mark2f = inst_1.query("CALC:MARK2:X?")
        mark2f = float(mark2f)
        mark2f = mark2f - mark1f
        x = ed_izm(mark2f, ' ГГц', ' МГц', ' кГц', ' Гц', ' мГц', ' мкГц', ' нГц')
        x.ed_izm()
        mark22.set(xz)
        #mark22.set("{0:.1f} Гц".format(mark2f))
        
        mark3 = inst_1.query("CALC:MARK3:Y?")
        mark3 = float(mark3)
        x = ed_izm(mark3, ' ГВ', ' МВ', ' кВ', ' В', ' мВ', ' мкВ', ' нВ')
        x.ed_izm()
        mark13.set(xz)
        #mark13.set("{0:.5f} мВ".format(mark3))
        mark3f = inst_1.query("CALC:MARK3:X?")
        mark3f = float(mark3f)
        mark3f = mark3f - mark1f
        x = ed_izm(mark3f, ' ГГц', ' МГц', ' кГц', ' Гц', ' мГц', ' мкГц', ' нГц')
        x.ed_izm()
        mark23.set(xz) 
        #mark23.set("{0:.1f} Гц".format(mark3f))

        kAM1 = (mark2 + mark3)/mark1*100
        kAM1 = float(kAM1)
        kAM.set("{0:.3f} %".format(kAM1))
                
        


root = Tk()
ix = (root.winfo_screenwidth() - root.winfo_reqwidth()) / 3
iy = (root.winfo_screenheight() - root.winfo_reqheight()) / 3
root.title("Modulations 1.1")
root.geometry("810x410+%d+%d" % (ix, iy))
root.iconbitmap("icon.ico")
root.resizable(width=False, height=False)
frame = Frame(root)
frame.pack()

ttk.Style().configure("TButton", padding=1, font='calibri 9', relief = "flat", foreground="black", background="cyan4")
ttk.Style().configure("TLabel", padding=1, font='calibri 9', relief = "flat", foreground="black")
#ttk.Style().configure("TEntry", padding=1, font='calibri 8', relief = "flat", foreground="black")

note = Notebook(root)
tab1 = Frame(note)
tab2 = Frame(note)
tab3 = Frame(note)
tab4 = Frame(note)
tab5 = Frame(note)
tab6 = Frame(note)

note.add(tab1, text = "  Program   ", compound=TOP)
note.add(tab2, text = "     AM     ")
note.add(tab3, text = "     FM     ")
note.add(tab4, text = "     ФМ     ")
note.add(tab5, text = "     IM     ")
note.add(tab6, text = "Information")
note.pack(fill=BOTH, expand=True)

today = datetime.datetime.today()
a=StringVar()
a.set('')
b=StringVar()
b.set('')
c=StringVar()
c.set('')
e=StringVar()



kAM=StringVar()
kAM.set('')
mark11=StringVar()
mark11.set('')
mark12=StringVar()
mark12.set('')
mark13=StringVar()
mark13.set('')
mark21=StringVar()
mark21.set('')
mark22=StringVar()
mark22.set('')
mark23=StringVar()
mark23.set('')


img = ImageTk.PhotoImage(Image.open("pan.png"))
imglabel = Label(tab1, image=img).grid(row=1, column=1)

but1 = ttk.Button(tab1, text = "Connect", command=click_connect)
but1.place(x=150,y=70)

#|||||||||||||||||||||||||||||||AM
but3 = ttk.Button(tab2, state=DISABLED, text = "Измерить", command=click_startAM)
but3.place(x=10,y=30)
lab1 = ttk.Label(tab2, text='Коэф.AM', style="TLabel", font='calibri 12')
lab1.place(x=125,y=5)
entry2 = ttk.Entry(tab2, state='readonly', textvariable = kAM, width = 8, font='calibri 20')
entry2.place(x=100,y=25)

combo2 = ttk.Combobox(tab2, values=span, state='readonly', height=10, width=6, font='calibri 12')
combo2.place(x=50,y=100)
lab11 = ttk.Label(tab2, text='Span', style="TLabel", font='calibri 12')
lab11.place(x=10,y=100)

lab3 = ttk.Label(tab2, text='Уровень несущей', style="TLabel", font='calibri 10')
lab3.place(x=250,y=25)
entry3 = ttk.Entry(tab2, state='readonly', textvariable = mark11, width = 10, font='calibri 12')
entry3.place(x=400,y=25)
lab4 = ttk.Label(tab2, text='Уровень первой боковой', style="TLabel", font='calibri 10')
lab4.place(x=250,y=50)
entry4 = ttk.Entry(tab2, state='readonly', textvariable = mark12, width = 10, font='calibri 12')
entry4.place(x=400,y=50)
lab5 = ttk.Label(tab2, text='Уровень второй боковой', style="TLabel", font='calibri 10')
lab5.place(x=250,y=75)
entry5 = ttk.Entry(tab2, state='readonly', textvariable = mark13, width = 10, font='calibri 12')
entry5.place(x=400,y=75)

lab6 = ttk.Label(tab2, text='Частота несущей', style="TLabel", font='calibri 10')
lab6.place(x=500,y=25)
entry6 = ttk.Entry(tab2, state='readonly', textvariable = mark21, width = 18, font='calibri 12')
entry6.place(x=650,y=25)
lab7 = ttk.Label(tab2, text='Частота первой боковой', style="TLabel", font='calibri 10')
lab7.place(x=500,y=50)
entry7 = ttk.Entry(tab2, state='readonly', textvariable = mark22, width = 14, font='calibri 12')
entry7.place(x=650,y=50)
lab8 = ttk.Label(tab2, text='Частота второй боковой', style="TLabel", font='calibri 10')
lab8.place(x=500,y=75)
entry8 = ttk.Entry(tab2, state='readonly', textvariable = mark23, width = 14, font='calibri 12')
entry8.place(x=650,y=75)




lab0 = ttk.Label(tab1, text='Port', style="TLabel")
lab0.place(x=10,y=55)
lab1 = ttk.Label(tab1, text='ID:', style="TLabel")
lab1.place(x=230,y=55)

lab9 = ttk.Label(tab6, text='Данная программа измеряет АМ, ЧМ, ФМ, ИМ с помощью:', style="TLabel")
lab9.place(x=20,y=10)
lab10 = ttk.Label(tab6, text='Анализатора спектра N9030A\n', style="TLabel")
lab10.place(x=180,y=30)

combo1 = ttk.Combobox(tab1, textvariable=e, state='readonly', height=10, width=19, font='calibri 9')
combo1.place(x=10,y=70)
entry1 = ttk.Entry(tab1, textvariable=a, state='readonly', width = 50, font='calibri 9')
entry1.place(x=230,y=70)


text=Text(tab1, state=DISABLED, width=40, height=20, wrap=WORD, font='calibri 9') #wrap=WORD ПЕРЕНОС ПО СЛОВАМ
text.place(x=542,y=65)
text.see("end")
scroll = Scrollbar(tab1, command=text.yview)
scroll.place(x=787,y=65, height=290)
text.yview('end')
text.config(yscrollcommand=scroll.set)

pb = ttk.Progressbar(root, orient=HORIZONTAL, length=785)
pb.place(x=10, y=380)
pb.config(mode="determinate", maximum=72, value=0)
pb1 = ttk.Progressbar(tab2, orient=HORIZONTAL, length=115)
pb1.place(x=102, y=70, height=10)
pb1.config(mode="determinate", maximum=5, value=0)

root.event_add('<<Paste>>', '<Control-igrave>')
root.event_add("<<Copy>>", "<Control-ntilde>")
tip()
root.mainloop()
